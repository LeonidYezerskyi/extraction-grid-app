"""Extract all unique sentiment labels from Excel file."""
import sys
import openpyxl
from collections import Counter

# Read the Excel file
file_path = 'Original_Final_Grid_Qual_Demo__GTM__01022026.xlsx'

try:
    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Available sheets:", wb.sheetnames)
    
    # Find sentiment sheet
    sentiment_sheet_name = None
    for name in wb.sheetnames:
        if 'sentiment' in name.lower():
            sentiment_sheet_name = name
            print(f"\nFound sentiment sheet: {sentiment_sheet_name}")
            break
    
    if not sentiment_sheet_name:
        print("ERROR: No sentiment sheet found!")
        sys.exit(1)
    
    # Get the sheet
    sheet = wb[sentiment_sheet_name]
    print(f"\nSheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
    
    # Collect all unique sentiment values
    all_sentiments = []
    
    # Go through all cells in the sheet
    for row in sheet.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                val_str = str(val).strip()
                if val_str and val_str.lower() not in ['nan', 'none', 'null', '']:
                    # Check if it's a numbered format (e.g., "1: positive")
                    # If so, extract just the sentiment part
                    if ':' in val_str:
                        parts = val_str.split(':', 1)
                        if len(parts) > 1:
                            sentiment_part = parts[1].strip()
                            if sentiment_part:
                                all_sentiments.append(sentiment_part)
                        # Also keep the full string
                        all_sentiments.append(val_str)
                    else:
                        all_sentiments.append(val_str)
    
    # Count occurrences
    sentiment_counts = Counter(all_sentiments)
    
    print(f"\nTotal sentiment values found: {len(all_sentiments)}")
    print(f"Unique sentiment values: {len(sentiment_counts)}")
    
    print("\n" + "=" * 80)
    print("ALL UNIQUE SENTIMENTS (sorted by frequency):")
    print("=" * 80)
    
    for sentiment, count in sentiment_counts.most_common():
        print(f"  {sentiment:40s} ({count} times)")
    
    print("\n" + "=" * 80)
    print("UNIQUE SENTIMENTS (alphabetically, for easy copy-paste):")
    print("=" * 80)
    
    unique_sorted = sorted(set(sentiment_counts.keys()))
    for sentiment in unique_sorted:
        # Clean up - remove numbered prefixes if any
        clean_sentiment = sentiment
        if ':' in clean_sentiment:
            parts = clean_sentiment.split(':', 1)
            if len(parts) > 1:
                clean_sentiment = parts[1].strip()
        
        # Remove common delimiters and split
        clean_sentiment = clean_sentiment.replace(';', ',').replace('\n', ' ')
        parts = [p.strip() for p in clean_sentiment.split(',') if p.strip()]
        
        for part in parts:
            if part and part.lower() not in ['nan', 'none', 'null', '']:
                print(f"  '{part}'")

except Exception as e:
    import traceback
    print(f"ERROR: {e}")
    traceback.print_exc()
